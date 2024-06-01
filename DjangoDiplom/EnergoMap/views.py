import os
from decimal import Decimal
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import load_workbook
from EnergoMap.models import Price, FDData, Subject


def welcome(request):
    return render(request, 'welcome.html')

def CFO(request):
    cfo_data = FDData.objects.get(id=1)
    cfo_subjects = Subject.objects.filter(ID_FD=1)
    cfo_prices = {}
    for subject in cfo_subjects:
        price = Price.objects.filter(id=subject.id).first()
        if price:
            cfo_prices[subject.id] = price.price
    context = {'cfo_subjects': cfo_subjects, 'cfo_prices': cfo_prices, 'cfo_data': cfo_data}
    return render(request, 'CFO.html', context)

def SZFO(request):
    szfo_data = FDData.objects.get(id=2)
    szfo_subjects = Subject.objects.filter(ID_FD=2)
    szfo_prices = {}
    for subject in szfo_subjects:
        price = Price.objects.filter(id=subject.id).first()
        if price:
            szfo_prices[subject.id] = price.price
    context = {'szfo_subjects': szfo_subjects, 'szfo_prices': szfo_prices, 'szfo_data': szfo_data}
    return render(request, 'SZFO.html', context)

def UFO(request):
    ufo_data = FDData.objects.get(id=3)
    ufo_subjects = Subject.objects.filter(ID_FD=3)
    ufo_prices = {}
    for subject in ufo_subjects:
        price = Price.objects.filter(id=subject.id).first()
        if price:
            ufo_prices[subject.id] = price.price
    context = {'ufo_subjects': ufo_subjects, 'ufo_prices': ufo_prices, 'ufo_data': ufo_data}
    return render(request, 'UFO.html', context)

def SKFO(request):
    skfo_data = FDData.objects.get(id=4)
    skfo_subjects = Subject.objects.filter(ID_FD=4)
    skfo_prices = {}
    for subject in skfo_subjects:
        price = Price.objects.filter(id=subject.id).first()
        if price:
            skfo_prices[subject.id] = price.price
    context = {'skfo_subjects': skfo_subjects, 'skfo_prices': skfo_prices, 'skfo_data': skfo_data}
    return render(request, 'SKFO.html', context)

def PFO(request):
    pfo_data = FDData.objects.get(id=5)
    pfo_subjects = Subject.objects.filter(ID_FD=5)
    pfo_prices = {}
    for subject in pfo_subjects:
        price = Price.objects.filter(id=subject.id).first()
        if price:
            pfo_prices[subject.id] = price.price
    context = {'pfo_subjects': pfo_subjects, 'pfo_prices': pfo_prices, 'pfo_data': pfo_data}
    return render(request, 'PFO.html', context)

def YFO(request):
    yfo_data = FDData.objects.get(id=6)
    yfo_subjects = Subject.objects.filter(ID_FD=6)
    yfo_prices = {}
    for subject in yfo_subjects:
        price = Price.objects.filter(id=subject.id).first()
        if price:
            yfo_prices[subject.id] = price.price
    context = {'yfo_subjects': yfo_subjects, 'yfo_prices': yfo_prices, 'yfo_data': yfo_data}
    return render(request, 'YFO.html', context)

def SFO(request):
    sfo_data = FDData.objects.get(id=7)
    sfo_subjects = Subject.objects.filter(ID_FD=7)
    sfo_prices = {}
    for subject in sfo_subjects:
        price = Price.objects.filter(id=subject.id).first()
        if price:
            sfo_prices[subject.id] = price.price
    context = {'sfo_subjects': sfo_subjects, 'sfo_prices': sfo_prices,'sfo_data': sfo_data}
    return render(request, 'SFO.html', context)

def DFO(request):
    dfo_data = FDData.objects.get(id=8)
    dfo_subjects = Subject.objects.filter(ID_FD=8)
    dfo_prices = {}
    for subject in dfo_subjects:
        price = Price.objects.filter(id=subject.id).first()
        if price:
            dfo_prices[subject.id] = price.price
    context = {'dfo_subjects': dfo_subjects, 'dfo_prices': dfo_prices,'dfo_data': dfo_data}
    return render(request, 'DFO.html', context)


#from EnergoMap.views import import_price_from_excel
#import_price_from_excel(None)
#from EnergoMap.views import import_fd_data_from_excel
#import_fd_data_from_excel(None)
#from EnergoMap.views import import_subjects_data_from_excel
#import_subjects_data_from_excel(None)

def import_price_from_excel(request):
    excel_file = os.path.abspath('static/excel/price.xlsx')
    wb = load_workbook(excel_file)
    ws = wb.active

    if Price.objects.exists():
        print("Данные уже загружены, повторная загрузка невозможна.")
        return
    for row in ws.iter_rows(min_row=2, values_only=True):
        price_value_str = str(row[0])  # Преобразуем значение в строку

        # Заменяем запятую на точку в строке
        price_value_str = price_value_str.replace(',', '.')

        try:
            # Преобразуем строку в десятичное число
            price_value = Decimal(price_value_str)

            price_instance = Price(price=price_value)
            price_instance.save()
            print(f"Price {price_value} saved successfully")
        except Exception as e:
            print(f"Error saving price {price_value_str}: {e}")
            continue

    return HttpResponse("Данные успешно импортированы")

def import_fd_data_from_excel(request):
    excel_file = os.path.abspath('static/excel/FD.xlsx')
    wb = load_workbook(excel_file)
    ws = wb.active

    if FDData.objects.exists():
        print("Данные уже загружены, повторная загрузка невозможна.")
        return
    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]  # Получаем название
        production = row[1]  # Получаем производство
        consumption = row[2]  # Получаем потребление

        try:
            fd_instance = FDData(name=name, production=production, consumption=consumption)
            fd_instance.save()
            print(f"Data {name}, Production {production}, Consumption {consumption} saved successfully")
        except Exception as e:
            print(f"Error saving data {name}, Production {production}, Consumption {consumption}: {e}")
            continue

    return HttpResponse("Данные успешно импортированы")

def import_subjects_data_from_excel(request):
    excel_file = os.path.abspath('static/excel/subjects.xlsx')
    wb = load_workbook(excel_file)
    ws = wb.active

    if Subject.objects.exists():
        print("Данные уже загружены, повторная загрузка невозможна.")
        return

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = row[0]  # Получаем название
        production = row[1]  # Получаем производство
        consumption = row[2]  # Получаем потребление
        id_fd = row[3]  # Получаем ID федерального округа

        try:
            subject_instance = Subject(name=name, production=production, consumption=consumption, ID_FD=id_fd)
            subject_instance.save()
            print(f"Data {name}, Production {production}, Consumption {consumption}, ID_FD {id_fd} saved successfully")
        except Exception as e:
            print(f"Error saving data {name}, Production {production}, Consumption {consumption}, ID_FD {id_fd}: {e}")
            continue

    return HttpResponse("Данные успешно импортированы")
def upload_file(request):
    if request.method == 'POST' and request.FILES['file']:
        uploaded_file = request.FILES['file']
        file_path = 'static/excel/' + uploaded_file.name


        # Проверяем, существует ли файл
        if os.path.exists(file_path):
            return HttpResponse('Файл с таким именем уже существует.')
        # Сохраняем файл на сервере
        with open('static/excel/' + uploaded_file.name, 'wb') as destination:
            for chunk in uploaded_file.chunks():
                destination.write(chunk)
        try:
            # Выполнение трех функций
            import_subjects_data_from_excel(request)
            import_fd_data_from_excel(request)
            import_price_from_excel(request)
        except Exception as e:
            print(f'Произошла ошибка при выполнении функций: {e}')

        return HttpResponse('Файл успешно загружен.')
